"""
Built-in reference documents fed to the AI on EVERY step so the analysis is
systematically aligned with IA2030 and the Gavi 6.0 (2026-2030) strategy.

- Important_Info_Tables_FR.xlsx: IA2030 priorities/indicators, Gavi 6.0 objectives &
  interventions (2026), aligned EPI components, main-intervention catalogue, scorecard.
- The EMR narrative template drives the Word report structure (see word_exporter), not the AI.
"""
from __future__ import annotations
from pathlib import Path

from config.settings import BASE_DIR
from core.models import UploadedDocument

REFERENCE_DIR = BASE_DIR / "reference_docs"

# Reference sheets, ordered SMALL CODE TABLES FIRST so exact codes (SPO / GIA / SPOGCInd) are
# always transmitted; the huge exhaustive catalogues come last and are abridged per-sheet.
KEY_SHEETS = [
    "IA2030 SP_SPO codes",          # 7 Strategic Priorities + 23 SPO codes (essential)
    "IA2030 aligned EPI comp",      # EPI component/subcomponent ↔ SPO ↔ Key Focus Areas
    "Gavi & IA2030",                # 8 Gavi investment domains (GIA 1-8) ↔ IA2030 SPA/SPO
    "IA2030 focus areas_Objs",      # IA2030 focus areas & objectives
    "NIS DevTeam",                  # roles/responsibilities (M&E responsible person)
    "IA2030SP_SPO_Indicators",      # SPOGCInd standard indicator catalogue
    "Main interventions",           # hierarchical intervention codification
    "Gavi 6.0 Obj-Intervtns 2026",  # Gavi objectives & intervention types (large -> abridged)
    "RegIA2030", "Scorecard",
]
_CAP = 27000            # total chars of reference text passed to the AI (cost control)
_PER_SHEET_CAP = 4000   # per-sheet char budget: keeps small code tables whole, abridges huge ones
_cache: list[UploadedDocument] | None = None


def _extract_key_sheets(path: Path, cap: int) -> str:
    import openpyxl
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    parts, total = [], 0
    names = [s for s in KEY_SHEETS if s in wb.sheetnames] or wb.sheetnames
    for name in names:
        if total >= cap:
            break
        ws = wb[name]
        header = f"\n=== {name} ==="
        parts.append(header); total += len(header)
        sheet_used = 0
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i > 200:
                break
            cells = [str(c).strip() for c in row if c is not None and str(c).strip()]
            if not cells:
                continue
            line = " | ".join(cells)
            if sheet_used + len(line) > _PER_SHEET_CAP:   # abridge this sheet, keep budget for the rest
                parts.append("… (onglet abrégé)")
                break
            parts.append(line); sheet_used += len(line); total += len(line)
            if total >= cap:
                parts.append("… (référence tronquée)")
                return "\n".join(parts)
    return "\n".join(parts)


def get_reference_documents() -> list[UploadedDocument]:
    """Return the bundled reference documents as UploadedDocuments (cached)."""
    global _cache
    if _cache is not None:
        return _cache
    docs: list[UploadedDocument] = []
    xlsx = REFERENCE_DIR / "Important_Info_Tables_FR.xlsx"
    if xlsx.exists():
        try:
            text = _extract_key_sheets(xlsx, _CAP)
            docs.append(UploadedDocument(
                name="Référence IA2030 & Gavi 6.0 — codes SPO, mapping EPI↔SPO, GIA 1-8, indicateurs SPOGCInd",
                file_type="xlsx", doc_category="Référence méthodologique OMS/IA2030/Gavi",
                text=text, tables_summary="[Codes normatifs: IA2030 SP/SPO, EPI↔SPO, Gavi GIA, SPOGCInd]"))
        except Exception:
            pass
    _cache = docs
    return docs
