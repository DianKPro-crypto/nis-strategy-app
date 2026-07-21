"""
Excel export — reproduces the structure of the original WHO workbook
"2_All in 1 SWOT to Activities_FR.xlsx".

Sheets produced:
  Country_Vision · Country_Sequence of events (Sections 1-5) ·
  SECTION 6_M&E · SECTION 7_Activities · How to prioritize
"""
from __future__ import annotations
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from config.settings import INSTITUTION_DARK, INSTITUTION_PRIMARY
from core.models import NISStrategy
from core.epi_components import EPI_COMPONENTS, subcomponent_pairs, find_subcomponent
from core.prioritization import INTERVENTION_CRITERIA
from core.branding import logo_path

_HDR = PatternFill("solid", fgColor=INSTITUTION_DARK.lstrip("#"))
_SUB = PatternFill("solid", fgColor=INSTITUTION_PRIMARY.lstrip("#"))
_COMP = PatternFill("solid", fgColor="DCE6F1")
_WHITE = Font(color="FFFFFF", bold=True)
_BOLD = Font(bold=True)
_WRAP = Alignment(wrap_text=True, vertical="top")
_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
_THIN = Border(*[Side(style="thin", color="BFBFBF")] * 4)


def _style_header(ws, row, ncols, fill=_HDR, font=_WHITE):
    for c in range(1, ncols + 1):
        cell = ws.cell(row, c)
        cell.fill = fill
        cell.font = font
        cell.alignment = _CENTER
        cell.border = _THIN


def _widths(ws, widths: dict[int, int]):
    for col, w in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = w


def build_excel(s: NISStrategy) -> bytes:
    lang = s.profile.language
    wb = Workbook()
    _sheet_vision(wb.active, s, lang)
    _sheet_sequence(wb.create_sheet("Country_Sequence of events"), s, lang)
    _sheet_me(wb.create_sheet("SECTION 6_M&E"), s, lang)
    _sheet_activities(wb.create_sheet("SECTION 7_Activities"), s, lang)
    _sheet_prioritize(wb.create_sheet("How to prioritize"), lang)
    _sheet_index(wb.create_sheet("Sommaire", 0), wb, s, lang)  # clickable table of contents
    wb.active = 0
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _sheet_index(ws, wb, s, lang):
    fr = lang == "fr"
    ws.sheet_view.showGridLines = False
    lp = logo_path(lang)
    if lp:
        try:
            from openpyxl.drawing.image import Image as XLImage
            img = XLImage(str(lp)); img.height, img.width = 50, int(50 * img.width / img.height)
            ws.add_image(img, "B2"); ws.row_dimensions[2].height = 42
        except Exception:
            pass
    period = f"{s.profile.nis_start_year}–{s.profile.nis_start_year + s.profile.nis_duration_years - 1}"
    ws["B5"] = ("Stratégie Nationale de Vaccination" if fr else "National Immunization Strategy")
    ws["B5"].font = Font(bold=True, size=18, color=INSTITUTION_DARK.lstrip("#"))
    ws["B6"] = f"{s.profile.country_name} · {period}"
    ws["B6"].font = Font(size=12, color=INSTITUTION_PRIMARY.lstrip("#"))
    ws["B8"] = ("Sommaire (cliquez pour accéder à chaque feuille)" if fr
                else "Table of contents (click to open each sheet)")
    ws["B8"].font = Font(bold=True, size=12)
    descriptions = {
        "Country_Vision": "Vision, but et objectif" if fr else "Vision, goal & objective",
        "Country_Sequence of events": "FFOM → causes → objectifs → interventions"
        if fr else "SWOT → causes → objectives → interventions",
        "SECTION 6_M&E": "Cadre de suivi & évaluation" if fr else "M&E framework",
        "SECTION 7_Activities": "Activités opérationnelles" if fr else "Operational activities",
        "How to prioritize": "Méthode de priorisation" if fr else "Prioritization method",
    }
    r = 10
    for i, name in enumerate([n for n in wb.sheetnames if n != "Sommaire"], 1):
        cell = ws.cell(r, 2, f"{i}.  {name}")
        cell.hyperlink = f"#'{name}'!A1"
        cell.font = Font(color="0563C1", underline="single", bold=True, size=12)
        ws.cell(r, 4, descriptions.get(name, "")).font = Font(italic=True, color="555555")
        ws.row_dimensions[r].height = 22
        r += 1
    ws.column_dimensions["A"].width = 3
    ws.column_dimensions["B"].width = 34
    ws.column_dimensions["C"].width = 3
    ws.column_dimensions["D"].width = 55


def _sheet_vision(ws, s, lang):
    ws.title = "Country_Vision"
    lp = logo_path(lang)
    if lp:
        try:
            from openpyxl.drawing.image import Image as XLImage
            img = XLImage(str(lp))
            img.height, img.width = 48, int(48 * img.width / img.height)
            ws.add_image(img, "A1")
            ws.row_dimensions[1].height = 38
        except Exception:
            pass
    ws["B1"] = "Vision, but et objectif général du pays"
    ws["B1"].font = _BOLD
    ws.append([])
    ws["B3"], ws["C3"], ws["D3"] = "Éléments", "Définition", f"Pays : {s.profile.country_name}"
    _style_header(ws, 3, 4)
    rows = [
        ("Vision du pays", "Énoncé de ce que le pays souhaite atteindre à long terme (≈10 ans).", s.vision.vision),
        ("But de la SNV", "Jalon intermédiaire sur la période de la SNV.", s.vision.goal),
        ("Objectif général", "Objectif général de la vision et du but du pays.", s.vision.overall_objective),
    ]
    r = 4
    for label, definition, value in rows:
        ws.cell(r, 2, label).font = _BOLD
        ws.cell(r, 3, definition).alignment = _WRAP
        ws.cell(r, 4, value or "").alignment = _WRAP
        r += 2
    _widths(ws, {2: 22, 3: 45, 4: 55})


def _sheet_sequence(ws, s, lang):
    """Sections 1-5 on one sheet, EPI rows in order."""
    headers = ["Composantes et sous-composantes du PEV", "Exemples de contenu à rechercher",
               "Forces (F)", "Faiblesse (Fa)", "Opportunités (O)", "Menaces (M)",
               "Faiblesse (Fa)", "POURQUOI-1", "POURQUOI-2", "POURQUOI-3", "Dernier POURQUOI",
               "Problème/Obstacle principal", "Résultat visionnaire du changement",
               "Objectif stratégique prioritaire", "Interventions principales",
               "Niveau de priorisation"]
    # Section banner row
    ws.cell(1, 1, "SECTION 1 : Analyse FFOM").font = _BOLD
    ws.cell(1, 7, "SECTION 2 : Causes profondes").font = _BOLD
    ws.cell(1, 12, "SECTION 3 : Obstacle").font = _BOLD
    ws.cell(1, 13, "SECTION 4 : Théorie du changement").font = _BOLD
    ws.cell(1, 15, "SECTION 5 : Interventions & priorisation").font = _BOLD
    for ci, h in enumerate(headers, 1):
        ws.cell(2, ci, h)
    _style_header(ws, 2, len(headers))

    swot_by_sub = {(x.component_code, x.subcomponent_code): x for x in s.swot}
    rc_by_sub: dict[str, list] = {}
    for rc in s.root_causes:
        rc_by_sub.setdefault(rc.subcomponent_code, []).append(rc)
    obj_by_sub: dict[str, list] = {}
    for o in s.objectives:
        obj_by_sub.setdefault(o.subcomponent_code, []).append(o)
    iv_by_obj: dict[str, list] = {}
    for iv in s.interventions:
        iv_by_obj.setdefault(iv.objective_id, []).append(iv)

    r = 3
    for comp in EPI_COMPONENTS:
        ws.cell(r, 1, comp.label(lang)).font = _BOLD
        for c in range(1, len(headers) + 1):
            ws.cell(r, c).fill = _COMP
        r += 1
        for sub in comp.subcomponents:
            sw = swot_by_sub.get((comp.code, sub.code))
            ws.cell(r, 1, sub.label(lang)).alignment = _WRAP
            ws.cell(r, 2, sub.examples_fr).alignment = _WRAP
            if sw:
                ws.cell(r, 3, "\n".join(sw.strengths)).alignment = _WRAP
                ws.cell(r, 4, "\n".join(sw.weaknesses)).alignment = _WRAP
                ws.cell(r, 5, "\n".join(sw.opportunities)).alignment = _WRAP
                ws.cell(r, 6, "\n".join(sw.threats)).alignment = _WRAP
            main_problem = ""
            for rc in rc_by_sub.get(sub.code, []):
                ws.cell(r, 7, rc.weakness).alignment = _WRAP
                for i, why in enumerate(rc.whys[:3]):
                    ws.cell(r, 8 + i, why).alignment = _WRAP
                ws.cell(r, 11, rc.final_why).alignment = _WRAP
                main_problem = main_problem or rc.main_problem
            # SECTION 3 — Problème principal (obstacle) : regroupement des derniers POURQUOI.
            if main_problem:
                ws.cell(r, 12, main_problem).alignment = _WRAP
            for o in obj_by_sub.get(sub.code, []):
                if not main_problem:
                    ws.cell(r, 12, o.main_obstacle).alignment = _WRAP
                ws.cell(r, 13, o.visionary_result).alignment = _WRAP
                ws.cell(r, 14, o.objective_text).alignment = _WRAP
                ivs = iv_by_obj.get(o.obj_id, [])
                ws.cell(r, 15, "\n".join(f"• {iv.title}" for iv in ivs)).alignment = _WRAP
                ws.cell(r, 16, "\n".join(_lvl(iv.priority_level, lang) for iv in ivs)).alignment = _WRAP
            r += 1
    _widths(ws, {1: 28, 2: 35, **{c: 18 for c in range(3, 17)}})
    ws.auto_filter.ref = f"A2:{get_column_letter(len(headers))}2"
    ws.freeze_panes = "C3"


def _sheet_me(ws, s, lang):
    ws.cell(1, 1, "SECTION 6 : Cadre de suivi et d’évaluation pour l’action").font = _BOLD
    base = ["Composante", "Sous-composante", "Objectif stratégique prioritaire", "Nom de l’indicateur",
            "Type", "Définition", "Calcul", "Source numérateur", "Source dénominateur",
            "Fréquence", "Responsable mesure", "Responsable action", "Situation de référence"]
    years = s.profile.years
    for ci, h in enumerate(base, 1):
        ws.cell(3, ci, h)
    ystart = len(base) + 1
    for i, y in enumerate(years):
        ws.cell(3, ystart + i, f"Cible {y}")
    ncols = ystart + len(years) - 1
    _style_header(ws, 3, ncols)
    r = 4
    for ind in s.indicators:
        cs = find_subcomponent(ind.subcomponent_code)
        comp_label = cs[0].label(lang) if cs else ind.component_code
        sub_label = cs[1].label(lang) if cs else ind.subcomponent_code
        vals = [comp_label, sub_label, ind.objective_id, ind.name, ind.indicator_type,
                ind.definition, ind.formula, ind.numerator_source, ind.denominator_source,
                ind.frequency, ind.responsible_measure, ind.responsible_action, ind.baseline]
        for ci, v in enumerate(vals, 1):
            ws.cell(r, ci, v).alignment = _WRAP
        for i, y in enumerate(years):
            ws.cell(r, ystart + i, ind.targets.get(f"Y{i+1}", "")).alignment = _WRAP
        r += 1
    _widths(ws, {1: 22, 2: 22, 3: 28, 4: 24, 6: 30, **{c: 16 for c in range(7, ncols + 1)}})
    ws.auto_filter.ref = f"A3:{get_column_letter(ncols)}3"
    ws.freeze_panes = "D4"


def _sheet_activities(ws, s, lang):
    ws.cell(1, 1, "SECTION 7 : Définition des activités clés et calendrier").font = _BOLD
    base = ["Composante", "Sous-composante", "Objectif stratégique prioritaire",
            "Intervention principale", "Activité clé", "Niveau de mise en œuvre"]
    years = s.profile.years
    for ci, h in enumerate(base, 1):
        ws.cell(3, ci, h)
    ystart = len(base) + 1
    for i, y in enumerate(years):
        ws.cell(3, ystart + i, f"An {y}")
    extra = ["Responsable", "Partenaires", "Prérequis", "Risques", "Livrables"]
    estart = ystart + len(years)
    for ci, h in enumerate(extra):
        ws.cell(3, estart + ci, h)
    ncols = estart + len(extra) - 1
    _style_header(ws, 3, ncols)

    iv_by_id = {iv.intervention_id: iv for iv in s.interventions}
    obj_by_id = {o.obj_id: o for o in s.objectives}
    r = 4
    for a in s.activities:
        cs = find_subcomponent(a.subcomponent_code)
        iv = iv_by_id.get(a.intervention_id)
        obj = obj_by_id.get(a.objective_id)
        vals = [cs[0].label(lang) if cs else a.component_code,
                cs[1].label(lang) if cs else a.subcomponent_code,
                obj.objective_text if obj else a.objective_id,
                iv.title if iv else a.intervention_id, a.activity, a.implementation_level]
        for ci, v in enumerate(vals, 1):
            ws.cell(r, ci, v).alignment = _WRAP
        for i in range(len(years)):
            ws.cell(r, ystart + i, "X" if a.years.get(f"Y{i+1}") else "").alignment = _CENTER
        for ci, v in enumerate([a.lead, ", ".join(a.partners), "; ".join(a.prerequisites),
                                "; ".join(a.risks), "; ".join(a.deliverables)]):
            ws.cell(r, estart + ci, v).alignment = _WRAP
        r += 1
    _widths(ws, {1: 20, 2: 20, 3: 28, 4: 26, 5: 32, 6: 22, **{c: 8 for c in range(ystart, ystart + len(years))}})
    ws.auto_filter.ref = f"A3:{get_column_letter(ncols)}3"
    ws.freeze_panes = "C4"


def _sheet_prioritize(ws, lang):
    ws.cell(1, 1, "Comment prioriser — Méthode 1 (notation multicritère)").font = _BOLD
    ws.cell(2, 1, "Score 3 = meilleur · 2 = modéré · 1 = faible · Total 8 critères: max 24, min 8").font = _BOLD
    header = ["Critère (intervention)", "Score 3", "Score 2", "Score 1"]
    for ci, h in enumerate(header, 1):
        ws.cell(4, ci, h)
    _style_header(ws, 4, 4)
    legend = {
        "expertise": ("Suffisante", "Modérée", "Aucune"),
        "return_on_investment": ("Élevé", "Modéré", "Faible"),
        "effectiveness": ("Élevé", "Modéré", "Faible"),
        "ease_of_implementation": ("Très facile", "Modérément facile", "Pas facile"),
        "negative_consequences": ("Aucune", "Légères", "Importantes"),
        "legal_constraints": ("Aucune", "Légères", "Importantes"),
        "health_system_impact": ("Élevé", "Modéré", "Faible"),
        "feasibility": ("Très faisable", "Modérément faisable", "Faible"),
    }
    r = 5
    for key, fr, en in INTERVENTION_CRITERIA:
        ws.cell(r, 1, fr if lang == "fr" else en)
        s3, s2, s1 = legend.get(key, ("", "", ""))
        ws.cell(r, 2, s3); ws.cell(r, 3, s2); ws.cell(r, 4, s1)
        r += 1
    r += 1
    ws.cell(r, 1, "Seuils: Élevé 17-24 · Moyen 9-16 · Faible 1-8").font = _BOLD
    r += 2
    ws.cell(r, 1, "Méthode 2 — Matrice 2x2 IMPACT × FAISABILITÉ").font = _BOLD
    r += 1
    grid = [["", "Faisabilité (+)", "Faisabilité (-)"],
            ["Impact (+)", "Impact élevé & très faisable (1)", "Impact élevé & moins faisable (2)"],
            ["Impact (-)", "Impact faible & très faisable (3)", "Impact faible & moins faisable (4)"]]
    for gr in grid:
        for ci, v in enumerate(gr, 1):
            ws.cell(r, ci, v).border = _THIN
        r += 1
    _widths(ws, {1: 40, 2: 28, 3: 28, 4: 18})


def _lvl(level, lang):
    m = {"high": ("Élevé", "High"), "medium": ("Moyen", "Medium"), "low": ("Faible", "Low")}
    v = getattr(level, "value", level)
    fr, en = m.get(v, (v, v))
    return fr if lang == "fr" else en
